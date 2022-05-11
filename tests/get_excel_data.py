
from openpyxl import load_workbook


execl_path = r'C:\AUTOMATION\PROJECT\data\data.xlsx'

workbook = load_workbook(execl_path)

worksheet = workbook.worksheets[0]

login_form_parameters = []
for i in range(2, worksheet.max_row+1):
    username = worksheet.cell(column=1,row=i).value
    password = worksheet.cell(column=2,row=i).value
    checkpoint = worksheet.cell(column=3,row=i).value
    tuple=username,password,checkpoint
    login_form_parameters.append(tuple)

